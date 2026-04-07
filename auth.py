##import tkinter as tk
##from openpyxl import load_workbook
##def register(name, city, password):
##    wb = load_workbook("students.xlsx")
##    ws = wb["Sheet1"]
##    if name == "" or city == "" or password == "":
##        return "All fields are required"
##    for row in ws.iter_rows(values_only=True):
##        if row[1] == name:
##            return "Already exists"
##    total = ws.max_row
##    new_id = name + "_" + str(total + 1)
##    ws.append([new_id, name, city, password, 0.0, 0])
##    wb.save("students.xlsx")
##    return "account created"
##
##def login(log_name, log_pass):
##    wb = load_workbook("students.xlsx")
##    ws = wb["Sheet1"]
##    for row in ws.iter_rows(values_only=True):
##        if row[1] == log_name:
##            if row[3] == log_pass:
##                return "login Successful", row[0]
##            else:
##                return "incorrect password", None
##    return "you are not registered", None
##
### =====================
### GUI FUNCTIONS
### =====================
##
##def open_register(root):
##    rw = tk.Toplevel(root)
##    rw.title("Register")
##    rw.geometry("300x250")
##
##    tk.Label(rw, text="Register", font=("Arial", 14, "bold")).pack(pady=10)
##    tk.Label(rw, text="Username").pack()
##    name_entry = tk.Entry(rw)
##    name_entry.pack()
##
##    tk.Label(rw, text="City").pack()
##    city_entry = tk.Entry(rw)
##    city_entry.pack()
##
##    tk.Label(rw, text="Password").pack()
##    pass_entry = tk.Entry(rw, show="*")
##    pass_entry.pack()
##
##    result = tk.Label(rw, text="")
##    result.pack()
##
##    def do_register():
##        name = name_entry.get()
##        city = city_entry.get()
##        password = pass_entry.get()
##        msg = register(name, city, password)
##        result.config(text=msg)
##        if msg == "account created":
##            rw.destroy()
##
##    tk.Button(rw, text="Register", command=do_register).pack(pady=10)
##
##
##def open_login(root, open_main_func):
##    lw = tk.Toplevel(root)
##    lw.title("Login")
##    lw.geometry("300x220")
##
##    tk.Label(lw, text="Login", font=("Arial", 14, "bold")).pack(pady=10)
##    tk.Label(lw, text="Username").pack()
##    name_entry = tk.Entry(lw)
##    name_entry.pack()
##
##    tk.Label(lw, text="Password").pack()
##    pass_entry = tk.Entry(lw, show="*")
##    pass_entry.pack()
##
##    result = tk.Label(lw, text="")
##    result.pack()
##
##    def do_login():
##        name = name_entry.get()
##        password = pass_entry.get()
##        msg, student_id = login(name, password)
##        result.config(text=msg)
##        if msg == "login Successful":
##            lw.destroy()
##            open_terms(root, student_id, open_main_func)
##
##    tk.Button(lw, text="Login", command=do_login).pack(pady=10)
##
##
##def open_terms(root, student_id, open_main_func):
##    tw = tk.Toplevel(root)
##    tw.title("Terms And Services")
##    tw.geometry("400x350")
##
##    tk.Label(tw, text="Welcome To SkillSwap", font=("Arial", 14, "bold")).pack(pady=10)
##    tk.Label(tw, text="SkillSwap - Students Skill Exchange Platform", font=("Arial", 11, "bold")).pack()
##    tk.Label(tw, text="A peer to peer skill exchange platform connecting\ncollege students to share knowledge without financial barriers", font=("Arial", 10)).pack(pady=5)
##    tk.Label(tw, text="No Money Needed — Exchange skills, not cash!", font=("Arial", 10, "bold")).pack(pady=5)
##    tb=tk.Text(tw,height=5,width=45)
##    tb.pack(padx=20,pady=5)
##    tt="1- Respect Each other - Be respectful with your exchange partner\n\n2- Honest Skills - Only add skills you genuinely know\n\n3- Complete Exchange - Always complete what you agreed to"
##    tb.insert(tk.END,tt)
##    tb.config(state="disabled")
####    tk.Label(tw, text="1- Respect Each other - Be respectful with your exchange partner", font=("Arial", 10)).pack(fill="x", padx=20)
####    tk.Label(tw, text="2- Honest Skills - Only add skills you genuinely know", font=("Arial", 10)).pack(fill="x", padx=20)
####    tk.Label(tw, text="3- Complete Exchange - Always complete what you agreed to", font=("Arial", 10)).pack(fill="x", padx=20)
##
##    agree = tk.BooleanVar()
##    tk.Checkbutton(tw, text="I agree to terms", variable=agree).pack(pady=10)
##
##    msg_label = tk.Label(tw, text="", fg="red")
##    msg_label.pack()
##
##    def next():
##        if agree.get():
##            tw.destroy()
##            open_main_func(root, student_id)
##        else:
##            msg_label.config(text="Please agree first!")
##
##    tk.Button(tw, text="Next ->", command=next).pack()
##
##
### =====================
### MAIN WINDOW
### =====================
##
##def start():
##    root = tk.Tk()
##    root.title("SkillSwap")
##    root.geometry("500x300")
##
##    tk.Label(root, text="Welcome to SkillSwap", font=("Arial", 16, "bold")).pack(pady=20)
##    tk.Label(root, text="Learn By Teaching - Teach By Learn", font=("Arial", 10)).pack()
##
##    from main import open_main
##
##    tk.Button(root, text="Login",
##              command=lambda: open_login(root, open_main)).pack(pady=10)
##    tk.Button(root, text="Register",
##              command=lambda: open_register(root)).pack(pady=5)
##
##    root.mainloop()
##
##start()
from PIL import Image, ImageTk, ImageDraw
from openpyxl import load_workbook
import tkinter as tk
import os
from main import open_main, skills_ui
from admin import open_admin_panel

# =====================
# BACKEND FUNCTIONS
# =====================
wb = load_workbook("students.xlsx")
ws = wb["Sheet1"]

def register(name, city, password):
    if name == "" or city == "" or password == "":
        return "All fields are required"
    already_exists = False
    for row in ws.iter_rows(values_only=True):
        if row[1] == name:
            already_exists = True
    if already_exists:
        return "already exists"
    else:
        total = ws.max_row
        new_id = name + "_" + str(total + 1)
        ws.append([new_id, name, city, password, 0.0, 0])
        wb.save("students.xlsx")
        return "account created"

def login(log_name, log_pass):
    wb = load_workbook("students.xlsx")
    ws = wb["Sheet1"]
    for row in ws.iter_rows(values_only=True):
        if row[1] == log_name:
            if row[3] == log_pass:
                return "login Successful", row[0]
            else:
                return "incorrect password", None
    return "you are not exists", None

# =====================
# GUI
# =====================
root = tk.Tk()
root.title("SkillSwap")
root.geometry("500x600")
root.configure(bg="#1a1a2e")

# Logo banao
img = Image.open(r"C:\Users\Indus\OneDrive\Desktop\files\logo.png.png")
img = img.resize((150, 150))
mask = Image.new("L", (150, 150), 0)
draw = ImageDraw.Draw(mask)
draw.ellipse((0, 0, 150, 150), fill=255)
img.putalpha(mask)
logo = ImageTk.PhotoImage(img)

# Global variables
login_btn = None
register_btn = None

def terms(root, student_id):
    tw = tk.Toplevel(root)
    tw.title("Terms And Services")
    tw.geometry("400x350")
    tw.configure(bg="#1a1a2e")
    tk.Label(tw, text="Welcome To SkillSwap",
             font=("Arial", 14, "bold"),
             bg="#1a1a2e", fg="white").pack(pady=10)
    tk.Label(tw, text="SkillSwap - Students Skill Exchange Platform",
             font=("Arial", 11, "bold"),
             bg="#1a1a2e", fg="#a0a0ff").pack()
    tk.Label(tw, text="A peer to peer skill exchange platform connecting\ncollege students to share knowledge without financial barriers",
             font=("Arial", 10),
             bg="#1a1a2e", fg="white").pack(pady=5)
    tk.Label(tw, text="No Money Needed — Exchange skills, not cash!",
             font=("Arial", 10, "bold"),
             bg="#1a1a2e", fg="#00ff88").pack(pady=5)
    tb = tk.Text(tw, height=5, width=45,
                 bg="#16213e", fg="white")
    tb.pack(padx=20, pady=5)
    tt = "1- Respect Each other\n\n2- Honest Skills\n\n3- Complete Exchange"
    tb.insert(tk.END, tt)
    tb.config(state="disabled")
    agree = tk.BooleanVar()
    tk.Checkbutton(tw, text="I agree to terms",
                   variable=agree,
                   bg="#1a1a2e", fg="white",
                   selectcolor="#16213e").pack(pady=10)
    msg_label = tk.Label(tw, text="",
                         fg="red", bg="#1a1a2e")
    msg_label.pack()
    def next():
        if agree.get():
            tw.destroy()
            open_main(root, student_id)
        else:
            msg_label.config(text="Please agree first!")
    tk.Button(tw, text="Next ->",
              bg="#4a4a8a", fg="white",
              relief="flat", command=next).pack()

def show_main():
    # Sab widgets hatao
    for w in root.winfo_children():
        w.destroy()

    global login_btn, register_btn

    # Logo
    tk.Label(root, image=logo,
             bg="#1a1a2e").pack(pady=10)
    root.logo = logo

    # Title
    tk.Label(root, text="SkillSwap",
             font=("Arial", 24, "bold"),
             bg="#1a1a2e", fg="white").pack()
    tk.Label(root,
             text="Learn By Teaching - Teach By Learn",
             font=("Arial", 10),
             bg="#1a1a2e", fg="#a0a0ff").pack(pady=5)
    tk.Frame(root, bg="#4a4a8a",
             height=2, width=300).pack(pady=10)

    # Buttons
##    login_btn = tk.Button(root, text="Login",
##                          bg="#4a4a8a", fg="white",
##                          font=("Arial", 12, "bold"),
##                          width=20, pady=8,
##                          relief="flat",
##                          command=check_login)
##    login_btn.pack(pady=8)
    selected = tk.IntVar()
    selected.set(1)

##    register_btn = tk.Button(root, text="Register",
##                             bg="#16213e", fg="white",
##                             font=("Arial", 12, "bold"),
##                             width=20, pady=8,
##                             relief="flat",
##                             command=open_register)
##    register_btn.pack(pady=5)

    tk.Label(root,
             text="Pakistan Ka Pehla Skill Exchange!",
             font=("Arial", 9),
             bg="#1a1a2e", fg="#00ff88").pack(pady=15)
    def on_select():
        if selected.get()==2:
            register_btn.config(state="disabled")
        else:
            register_btn.config(state="normal")
##    btn_frame=tk.Frame(root,bg="#1a1a2e")
##    btn_frame.pack(pady=10)
##    selected = tk.IntVar()
##    selected.set(1)
##
##    tk.Radiobutton(root,text="Student",variable=selected,value=1,bg="#1a1a2e",fg="white").pack()
##    tk.Radiobutton(root,text="Admin",variable=selected,value=2,bg="#1a1a2e",fg="#ffd700").pack()
    def check_login():
        if selected.get() == 1:
            open_login()
        else:
            open_admin_login()
    login_btn = tk.Button(root, text="Login",
                          bg="#4a4a8a", fg="white",
                          font=("Arial", 12, "bold"),
                          width=20, pady=8,
                          relief="flat",
                          command=check_login)
    login_btn.pack(pady=8)
    register_btn = tk.Button(root, text="Register",
                             bg="#16213e", fg="white",
                             font=("Arial", 12, "bold"),
                             width=20, pady=8,
                             relief="flat",
                             command=open_register)
    register_btn.pack(pady=5)
##        selected = tk.IntVar()
##        selected.set(1)

    tk.Radiobutton(root,text="Student",variable=selected,value=1,command=on_select,bg="#1a1a2e",fg="white",selectcolor="#4a4a8a",activebackground="#1a1a2e",activeforeground="white").pack()
    tk.Radiobutton(root,text="Admin",variable=selected,value=2,command=on_select,bg="#1a1a2e",fg="#ffd700",selectcolor="#4a4a8a",activebackground="#1a1a2e",activeforeground="#ffd700").pack()
##    if selected.get()==1:
##        open_login()
##    else:
##        open_admin_login()
    
##    tk.Button(btn_frame,text="👨‍🎓 Student",bg="#16213e",fg="#a0a0ff",font=("Arial",10),width=12,pady=5,relief="flat",command=open_login).pack(side="left",padx=10)
##    tk.Button(btn_frame,text="👑 Admin",bg="#16213e",fg="#ffd700",font=("Arial",10),width=12,pady=5,relief="flat",command=open_admin_login).pack(side="left",padx=10)
def open_admin_login():
    global login_btn, register_btn
    for w in root.winfo_children():
        w.destroy()
    tk.Label(root,image=logo,bg="#1a1a2e").pack(pady=10)
    root.logo=logo
    tk.Label(root,text="👑 Admin Login",font=("Arial",18,"bold"),bg="#1a1a2e",fg="#ffd700").pack(pady=10)
##    login_btn.pack_forget()
##    register_btn.pack_forget()

    tk.Label(root, text="Username",
             bg="#1a1a2e", fg="#a0a0ff",font=("Arial",10)).pack()
    name_entry = tk.Entry(root)
    name_entry.pack()
    tk.Label(root, text="Password",
             bg="#1a1a2e", fg="#a0a0ff",font=("Arial",10)).pack()
    pass_entry = tk.Entry(root, show="*")
    pass_entry.pack()
    result = tk.Label(root, text="",
                      bg="#1a1a2e", fg="white")
    result.pack()

    def do_admin_login():
        name = name_entry.get()
        password = pass_entry.get()
        if name == "yahya" and password == "yahyatalha":
##        msg, student_id = login(name, password)
            result.config(text="Welcome Admin!")
            open_admin_panel(root)
        else:
            result.config(text="Wrong credentials!")
##        if msg == "login Successful":
##            terms(root, student_id)

    tk.Button(root, text="Login",
              bg="#4a4a8a", fg="white",font=("Arial",12,"bold"),width=20,pady=8,relief="flat",
              command=do_admin_login).pack(pady=8)

    def go_back():
        show_main()

    tk.Button(root, text="Back",
              bg="#16213e", fg="#a0a0ff",font=("Arial",12,"bold"),width=20,pady=8,relief="flat",
              command=go_back).pack(pady=5)
    

def open_login():
    global login_btn, register_btn
##    login_btn.pack_forget()
##    register_btn.pack_forget()
    for w in root.winfo_children():
        w.destroy()
    tk.Label(root,image=logo,bg="#1a1a2e").pack(pady=10)
    root.logo=logo
    tk.Label(root,text="👨‍🎓 Student Login",font=("Arial",18,"bold"),bg="#1a1a2e",fg="#00ff88").pack(pady=10)
    

    tk.Label(root, text="Username",
             bg="#1a1a2e", fg="#a0a0ff",font=("Arial",10)).pack()
    name_entry = tk.Entry(root)
    name_entry.pack()
    tk.Label(root, text="Password",
             bg="#1a1a2e", fg="#a0a0ff",font=("Arial",10)).pack()
    pass_entry = tk.Entry(root, show="*")
    pass_entry.pack()
    result = tk.Label(root, text="",
                      bg="#1a1a2e", fg="white")
    result.pack()

    def do_login():
        name = name_entry.get()
        password = pass_entry.get()
        msg, student_id = login(name, password)
        result.config(text=msg)
        if msg == "login Successful":
            terms(root, student_id)

    tk.Button(root, text="Login",
              bg="#4a4a8a", fg="white",font=("Arial",12,"bold"),width=20,pady=8,relief="flat",
              command=do_login).pack(pady=8)

    def go_back():
        show_main()

    tk.Button(root, text="Back",
              bg="#16213e", fg="#a0a0ff",font=("Arial",12,"bold"),width=20,pady=8,relief="flat",
              command=go_back).pack(pady=5)

def open_register():
    global login_btn, register_btn
##    login_btn.pack_forget()
##    register_btn.pack_forget()
    for w in root.winfo_children():
        w.destroy()
    tk.Label(root,image=logo,bg="#1a1a2e").pack(pady=10)
    root.logo=logo
    tk.Label(root,text="👨‍🎓 Student Register",font=("Arial",18,"bold"),bg="#1a1a2e",fg="#00ff88").pack(pady=10)

    tk.Label(root, text="Username",
             bg="#1a1a2e", fg="#a0a0ff",font=("Arial",10)).pack()
    name_entry = tk.Entry(root)
    name_entry.pack()
    tk.Label(root, text="City",
             bg="#1a1a2e", fg="#a0a0ff",font=("Arial",10)).pack()
    city_entry = tk.Entry(root)
    city_entry.pack()
    tk.Label(root, text="Password",
             bg="#1a1a2e", fg="#a0a0ff",font=("Arial",10)).pack()
    pass_entry = tk.Entry(root, show="*")
    pass_entry.pack()
    result = tk.Label(root, text="",
                      bg="#1a1a2e", fg="white")
    result.pack()

    def do_register():
        name = name_entry.get()
        city = city_entry.get()
        password = pass_entry.get()
        msg = register(name, city, password)
        result.config(text=msg)
        if msg == "account created":
            go_back()

    tk.Button(root, text="Register",
              bg="#4a4a8a", fg="white",font=("Arial",12,"bold"),width=20,pady=8,relief="flat",
              command=do_register).pack(pady=8)

    def go_back():
        show_main()

    tk.Button(root, text="Back",
              bg="#16213e", fg="#a0a0ff",font=("Arial",12,"bold"),width=20,pady=8,relief="flat",
              command=go_back).pack(pady=5)

# =====================
# START
# =====================
show_main()  # ← sab yahan se shuru!
root.mainloop()

