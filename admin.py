import tkinter as tk
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import datetime
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
def gt_analytics():
    wb = load_workbook("students.xlsx")
    ws = wb["Sheet1"]
    total = ws.max_row - 1
    ws2 = wb["Skills"]
    active = ws2.max_row - 1
    inactive = total - active
    msg_count = {}
    ws3 = wb["message"]
    for row in ws3.iter_rows(values_only = True,min_row = 2):
        if row[0] not in msg_count:
            msg_count[row[0]] = 0
        msg_count[row[0]] += 1
    top_messenger = ""
    max_msgs = 0
    for sid in msg_count:
        if msg_count[sid] > max_msgs:
            max_msgs = msg_count[sid]
            top_messenger = sid
    return total, active, inactive, top_messenger
        
def gt_data():
    wb = load_workbook("students.xlsx")
    ws1 = wb["Sheet1"]
    students = ws1.max_row - 1
    total_req = 0
    accepted = 0
    if "Requests" in wb.sheetnames:
        ws2 = wb["Requests"]
        for row in ws2.iter_rows(values_only=True):
            if row[0] != "from_id":
                total_req +=1
                if row[2] == "accepted":
                    accepted +=1
    messages = 0
    if "message" in wb.sheetnames:
        ws3 = wb["message"]
        messages = ws3.max_row - 1
    return students,total_req,accepted,messages
            
def open_admin_panel(root):
    admin_win = tk.Toplevel(root)
    admin_win.title("SkillSwap - Admin Panel")
    admin_win.state("zoomed")
    admin_win.configure(bg="#1a1a2e")

    header = tk.Frame(admin_win,bg="#0f0f23")
    header.pack(fill="x")
    tk.Label(header,text="👑 SkillSwap Admin Panel",font=("Arial",20,"bold"),bg="#0f0f23",fg="#ffd700").pack(pady=10)
    tk.Label(header,text="Welcome Yahya",font=("Arial",11),bg="#0f0f23",fg="#a0a0ff").pack(side="right",padx=20)

    navbar = tk.Frame(admin_win,bg="#16213e")
    navbar.pack(fill="x")
    content = tk.Frame(admin_win,bg="#1a1a2e")
    content.pack(fill="both",expand=True)
    def clear():
        for w in content.winfo_children():
            w.destroy()
    def make_card(parent,title,value,color):
        card = tk.Frame(parent,bg="#16213e",relief="raised",bd=2)
        card.pack(side="left",padx=20,pady=10,ipadx=30,ipady=20)
        tk.Label(card,text=title,bg="#16213e",fg="#a0a0ff",font=("Arial",12)).pack(pady=5)
        tk.Label(card,text=str(value),bg="#16213e",fg=color,font=("Arial",36,"bold")).pack(pady=5)
        
    def dashboard():
##        pass
        clear()
        students,total_req,accepted,msgs = gt_data()
        tk.Label(content,text="📊 Live Dashboard",font=("Arial",16,"bold"),bg="#1a1a2e",fg="#ffd700").pack(pady=15)
        cards = tk.Frame(content,bg="#1a1a2e")
        cards.pack(pady=20)
        def make_card(parent,title,value,color):
            card = tk.Frame(parent,bg="#16213e",relief="raised",bd=2)
            card.pack(side="left",padx=20,pady=10,ipadx=30,ipady=20)
            tk.Label(card,text=title,bg="#16213e",fg="#a0a0ff",font=("Arial",12)).pack(pady=5)
            tk.Label(card,text=str(value),bg="#16213e",fg=color,font=("Arial",36,"bold")).pack(pady=5)
        make_card(cards, "👥 Total Students",
              students, "#00ff88")
        make_card(cards, "📋 Total Requests",
              total_req, "#a0a0ff")
        make_card(cards, "✅ Accepted",
              accepted, "#ffd700")
        make_card(cards, "💬 Messages",
              msgs, "#ff6b6b")
            
        
    def reports():
        clear()
        tk.Label(content, text="Students Report", font=("Arial", 16, "bold"), bg="#1a1a2e", fg="#ffd700").pack(pady=15)
        wb = load_workbook("students.xlsx")
        ws = wb["Sheet1"]

        header_row = tk.Frame(content, bg="#0f0f23")
        header_row.pack(fill="x", padx=10)
        tk.Label(header_row, text="ID",       bg="#0f0f23", fg="#ffd700", font=("Arial", 11, "bold"), width=15).pack(side="left")
        tk.Label(header_row, text="Name",     bg="#0f0f23", fg="#ffd700", font=("Arial", 11, "bold"), width=15).pack(side="left")
        tk.Label(header_row, text="City",     bg="#0f0f23", fg="#ffd700", font=("Arial", 11, "bold"), width=15).pack(side="left")
        tk.Label(header_row, text="Ratings",  bg="#0f0f23", fg="#ffd700", font=("Arial", 11, "bold"), width=15).pack(side="left")
        tk.Label(header_row, text="Exchange", bg="#0f0f23", fg="#ffd700", font=("Arial", 11, "bold"), width=15).pack(side="left")
        tk.Label(header_row, text="Actions",  bg="#0f0f23", fg="#ffd700", font=("Arial", 11, "bold"), width=15).pack(side="left")

    # --- Load admin_reports ek baar ---
        wb2 = load_workbook("students.xlsx")
        msg_data = {}
        if "admin_reports" in wb2.sheetnames:
            ws_rep = wb2["admin_reports"]
            for r in ws_rep.iter_rows(values_only=True, min_row=2):
                sid = r[0]
                if sid not in msg_data:
                    msg_data[sid] = []
                msg_data[sid].append(r)

        for i, row in enumerate(ws.iter_rows(values_only=True, min_row=2)):
            bg = "#16213e" if i % 2 == 0 else "#0f0f23"
            row_frame = tk.Frame(content, bg=bg)
            row_frame.pack(fill="x", padx=10)

            tk.Label(row_frame, text=row[0], bg=bg, fg="white", width=19).pack(side="left")
            tk.Label(row_frame, text=row[1], bg=bg, fg="white", width=19).pack(side="left")
            tk.Label(row_frame, text=row[2], bg=bg, fg="white", width=19).pack(side="left")
            tk.Label(row_frame, text=row[4], bg=bg, fg="white", width=19).pack(side="left")
            tk.Label(row_frame, text=row[5], bg=bg, fg="white", width=19).pack(side="left")

        # --- Delete button ---
            def delete(sid=row[0]):
                wb3 = load_workbook("students.xlsx")
                ws3 = wb3["Sheet1"]
                for r2 in ws3.iter_rows():
                    if r2[0].value == sid:
                        ws3.delete_rows(r2[0].row)
                        wb3.save("students.xlsx")
                        break
                reports()

            tk.Button(row_frame, text="🗑 Delete", bg="red", fg="white",
                  relief="flat", command=delete).pack(side="left", padx=3)

        # --- Messages button (sirf tab dikhe jab message ho) ---
            student_msgs = msg_data.get(row[0], [])
            msg_count = len(student_msgs)

            def open_msg_popup(msgs=student_msgs, sname=row[1], sid=row[0]):
                popup = tk.Toplevel()
                popup.title("Messages - " + sname)
                popup.geometry("450x350")
                popup.configure(bg="#1a1a2e")
                popup.resizable(False, False)

                tk.Label(popup,
                     text="📢 " + sname + " ke Admin Messages",
                     font=("Arial", 13, "bold"),
                     bg="#1a1a2e", fg="#ffd700").pack(pady=12)

                if len(msgs) == 0:
                    tk.Label(popup, text="Koi message nahi!",
                         bg="#1a1a2e", fg="#a0a0ff").pack(pady=20)
                else:
                    for m in msgs:
                        card = tk.Frame(popup, bg="#16213e", relief="raised", bd=1)
                        card.pack(fill="x", padx=20, pady=4)
                        tk.Label(card,
                             text="📅 " + str(m[2]) + "   🕐 " + str(m[3]),
                             bg="#16213e", fg="#a0a0ff",
                             font=("Arial", 9)).pack(anchor="e", padx=8)
                        tk.Label(card,
                             text="💬 " + str(m[1]),
                             bg="#16213e", fg="white",
                             font=("Arial", 11)).pack(anchor="w", padx=8, pady=4)

                tk.Button(popup, text="Close",
                      bg="#4a4a8a", fg="white",
                      relief="flat", font=("Arial", 11),
                      command=popup.destroy).pack(pady=10)

        # Badge color — red agar message hai, grey agar nahi
            if msg_count > 0:
                btn_text = f"💬 {msg_count} Msg"
                btn_color = "#e63946"
            else:
                btn_text = "💬 0 Msg"
                btn_color = "#3a3a5c"

            tk.Button(row_frame, text=btn_text,
                  bg=btn_color, fg="white",
                  relief="flat", font=("Arial", 9),
                  command=open_msg_popup).pack(side="left", padx=3)
            

            
##        footer_row = tk.Frame(content,bg="#16213e")
##        footer_row.pack(fill="x",expand=True)
##        for row in ws.iter_rows(values_only=True):
            
            
    
    def anouncements():
        clear()
##        pass
        tk.Label(content,text="Announcements",font=("Arial",16,"bold"),bg="#1a1a2e",fg="#ffd700").pack(pady=15)
        msg_frame = tk.Frame(content,bg="#16213e")
        msg_frame.pack()
        msg_box = tk.Text(msg_frame,height = 4,width = 50,bg="#0f0f23",fg="white",font=("Arial",11))
        msg_box.pack(padx=10,pady=10)
        def send():
            msg = msg_box.get("1.0","end").strip()
            if msg == "":
                return "error"
            else:
                wb = load_workbook("students.xlsx")
                ws = wb["announcements"]
                now = datetime.datetime.now()
                date = now.strftime("%d-%m-%Y")
                time = now.strftime("%H:%M")
                ws.append([msg,date,time])
                wb.save("students.xlsx")
                msg_box.delete("1.0","end")
                anouncements()
        tk.Button(content,text="📢 Send",command=send).pack(pady=5)
        tk.Label(content,text="Old Announcements",bg="#1a1a2e",fg="#ffd700",font=("Arial",14)).pack(pady=15)
        wb = load_workbook("students.xlsx")
        ws = wb["announcements"]
        for r in ws.iter_rows(values_only=True, min_row=2):
            card = tk.Frame(content,bg="#16213e",relief="raised",bd=1)
            card.pack(fill="x",padx=100,pady=5)
            
##            row[0] = message
##            row[1] = date
##            row[2] = time
            tk.Label(card,text="📅 " + str(r[1]) + " " + "  🕐 " + str(r[2]), bg="#16213e",fg="#a0a0ff",font=("Arial",9)).pack(anchor="e",padx=10)
            tk.Label(card,text="📢 " + str(r[0]),bg="#16213e",fg="white",font=("Arial",11)).pack(anchor = "w",padx=10,pady=5)
##            tk.Label(content,text=r[1]).pack()
##            tk.Label(content,text=r[2]).pack()
                
        
    def platform_analytics():
        clear()
        total, active, inactive, top = gt_analytics()
        wb = load_workbook("students.xlsx")
        ws = wb["Skills"]
        names = []
        counts = []
        for row in ws.iter_rows(values_only = True, min_row = 2):
            if row[1]:
                names.append(row[0])
                skills = row[1].split("-")
                counts.append(len(skills))
        tk.Label(content,text="📈 Platform Analytics",font=("Arial",16,"bold"),bg="#1a1a2e",fg="#ffd700").pack(pady = 15)
       
        cards = tk.Frame(content,bg="#1a1a2e")
        cards.pack(pady=20)
##        tk.Label(card,text=title,bg="#16213e",fg="#a0a0ff",font=("Arial",12)).pack(pady=5)
##        tk.Label(card,text=str(value),bg="#16213e",fg=color,font=("Arial",36,"bold")).pack(pady=5)
        make_card(cards, "👥 Total Students",
              total, "#00ff88")
        make_card(cards, "📋 Active",
              active, "#a0a0ff")
        make_card(cards, "😴 Inactive",
              inactive, "#ffd700")
        make_card(cards, "💬 Top Messenger",
              top, "#ff6b6b")
        plt.close("all")
        fig, (ax1, ax2) = plt.subplots(1,2)
        fig.patch.set_facecolor("#1a1a2e")

        ax1.bar(names, counts, color="#4a4a8a")
        ax1.set_title("Student Skills Count",color="white")
        ax1.tick_params(colors="white",axis = "x",rotation = 45)
        ax1.tick_params(colors="white",axis="y")
        ax1.set_facecolor("#16213e")
        ax2.pie([active, inactive],labels=["Active", "Inactive"],colors=["#00ff88", "#ff6b6b"],autopct="%1.1f%%",textprops={"color": "white"})
        ax2.set_title("Active vs Inactive",color="white")
        ax2.set_facecolor("#16213e")

    # Tkinter mein embed
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=content)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True,padx=10, pady=5)
##        pass
    def search():
        clear()
        tk.Label(content,text="🔍 Search Student",font=("Arial",16,"bold"),bg="#1a1a2e",fg="#ffd700").pack(pady=15)
        search_frame = tk.Frame(content,bg="#1a1a2e")
        search_frame.pack(pady=10)
        
        search_entry = tk.Entry(search_frame,width=30,bg="#16213e",fg="white",font=("Arial",12),insertbackground="white")
        search_entry.pack(side = "left",padx=5)
##        tk.Button(content,text="🔍 Search",bg="#4a4a8a",fg="teal",relief="flat",command=do_search).pack(side="left")
        result_frm = tk.Frame(content,bg="#1a1a2e")
        result_frm.pack(fill="both",expand=True,pady=10)
        def do_search():
            for w in result_frm.winfo_children():
                w.destroy()
            name = search_entry.get().lower()
            wb=load_workbook("students.xlsx")
            ws = wb["Sheet1"]
            found = False
            for row in ws.iter_rows(values_only=True,min_row=2):
                if name in str(row[1]).lower():
                    found = True
                    def show(r=row):
                        show_profile(r)
                    tk.Button(result_frm,text="👤 " + str(row[1]) + "   🏙️ " + str(row[2]),bg="#16213e",fg="white",relief="flat",font=("Arial",11),command=show).pack(fill="x",padx=50,pady=3)
            if not found:
                tk.Label(result_frm,text="No Student found!",bg="#1a1a2e",fg="red").pack()
        tk.Button(search_frame,text="🔍 Search",bg="#4a4a8a",fg="teal",relief="flat",font=("Arial",12),padx=10,command=do_search).pack(side="left",padx=5)

        def show_profile(row):
            for w in result_frm.winfo_children():
                w.destroy()
            student_id = row[0]
            name = row[1]
            city = row[2]
            ratings = row[4]

            wb = load_workbook("students.xlsx")
            ws = wb["Skills"]
            teach = "Not Added"
            learn = "Not Added"
            for r in ws.iter_rows(values_only=True):
                if r[0] == student_id:
                    teach = str(r[1]).replace("-", ",")
                    learn = str(r[2]).replace("-", ",")
            card = tk.Frame(result_frm,bg="#16213e",relief="raised",bd=4)
            card.pack(fill="x",padx=50,pady=10)
            tk.Label(card,text="👤 " + str(name),font=("Arial",18,"bold"),bg="#16213e",fg="#00ff88").pack(pady=10)
            tk.Label(card,text="🆔 " + str(student_id),bg="#16213e",fg="white",font=("Arial",11)).pack()
            tk.Label(card,text="🏙️ City: " + str(city),bg="#16213e", fg="white",font=("Arial", 11)).pack()
        
            tk.Label(card,text="⭐ Ratings: " + str(ratings),bg="#16213e", fg="#ffd700",font=("Arial", 11)).pack()
        
            tk.Label(card,text="🎓 Teaching: " + teach,bg="#16213e", fg="#00ff88",font=("Arial", 11)).pack(pady=5)
        
            tk.Label(card,text="📚 Learning: " + learn,bg="#16213e", fg="#00ff88",font=("Arial", 11)).pack(pady=(0,10))
        
        # Back button
            tk.Button(result_frm,text="← Back to Search",bg="#4a4a8a", fg="white",relief="flat",command=do_search).pack(pady=10)
            
                               
        
##        pass
    pages=[
        ("📊 Dashboard",dashboard),
        ("📋 Reports",reports),
        ("📢 Anouncements",anouncements),
        ("🏆 Platform Analytics",platform_analytics),
        ("🔍 Search Students",search)
        ]
    for text, cmd in pages:
        tk.Button(navbar,text=text,bg="#4a4a8a",fg="white",relief="flat",padx=15,pady=5,command=cmd).pack(side="left",expand=True,fill="x",padx=2)
        
    tk.Button(navbar,text="🚪 Logout",bg="red",fg="white",relief="flat",padx=15,pady=5,command=admin_win.destroy).pack(side="right",padx=5)
            
                        
