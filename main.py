import tkinter as tk
from openpyxl import load_workbook
import datetime
from PIL import Image,ImageTk,ImageDraw
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

##def message():
##    wb=load_workbook("students.xlsx")
##    ws=wb["message"]
##    from_id = input("Enter from id: ")
##    to_id = input("Enter to_id: ")
##    if from_id == "" or to_id == "":
##        print("Enter both values")
##    else:
##        msg = input("Enter message: ")
##        row[0]=from_id
##        row[1]=to_id
##        row[2]=message
##message()
    

def save_skills(student_id, teaching, learning):
    wb = load_workbook("students.xlsx")
    ws = wb["Skills"]
    for row in ws.iter_rows():
        if row[0].value == student_id:
            row[1].value = teaching
            row[2].value = learning
            wb.save("students.xlsx")
            return
    ws.append([student_id, teaching, learning])
    wb.save("students.xlsx")
def request(from_id,to_id,common):
    wb = load_workbook("students.xlsx")
    if "Requests" not in wb.sheetnames:
        ws = wb.create_sheet("Requests")
        ws.append(["from_id","to_id","status","common_skills"])
    else:
        ws = wb["Requests"]
    for row in ws.iter_rows(values_only=True):
        if row[0]== from_id and row[1] == to_id:
##            tk.Label(content,text="Already Sent").pack(side="right",pady=5)
            return"Already Sent"
    com_st = "-".join(common)
    ws.append([from_id,to_id,"pending",com_st])
    wb.save("students.xlsx")
    return"sent"


categories={"Computer 💻": ["Python","JAVA","PHP","SQL","C#",
                             "C++","JavaScript","React","HTML","CSS","WordPress",
                             "PhotoShop","Video Editing"],
             
                "Cooking 🍳": ["Baking","BBQ","Biryani","Cake Decoration",
                            "Pizza","Pakistani Dishes","Chinese Food",
                            "Desserts"],
                "Arts 🎨": ["Drawing","Painting","Calligraphy","Sketching","Digital Art",
                         "Origami","Pottery"],
                "Singing 🎵": ["Classical","Pop","Nasheed","Rap","Folk","Qawali"],
                "Languages 🌍": ["English","Urdu","Arabic","French","Chinese",
                              "German"],
                "Academic 📖": ["Math","Physics","Chemistry","Biology","History",
                             "Islamiat"],
                "Fitness 💪": ["Yoga","Gym","Martial Arts","Meditation"]}
def skills_ui(student_id,frame):
    tk.Label(frame,text="Add Your Skills",font=("Arial",14,"bold"),bg="#1a1a2e",fg="white").pack(pady=10)
    teach_vars={}
    learn_vars={}
    for cat in categories:
        for skill in categories[cat]:
            teach_vars[skill]=tk.BooleanVar()
            learn_vars[skill]=tk.BooleanVar()

    cat_frame = tk.Frame(frame,bg="#1a1a2e")
    cat_frame.pack(pady=10)
    left=tk.Frame(frame,bg="#1a1a2e")
    left.pack(side="left",padx=50)
    right=tk.Frame(frame,bg="#1a1a2e")
    right.pack(side="right",padx=50)
    instruct = tk.Label(frame,text="Uper Se Category Select Karo!",font=("Arial",12),bg="#1a1a2e",fg="#a0a0ff")
    instruct.pack(pady=50)
    tk.Label(left,text="🎓 Jo Mujhe Ata Hy",bg="#1a1a2e",fg="white",font=("Arial",11,"bold")).pack()
    tk.Label(right,text="📚 Jo Mujhe Seekhna Hy",bg="#1a1a2e",fg="white",font=("Arial",11,"bold")).pack()
    def show_skills(cat):
        instruct.pack_forget()
        for w in left.winfo_children()[1:]:
            w.destroy()
        for w in right.winfo_children()[1:]:
            w.destroy()
        for skill in categories[cat]:
            tk.Checkbutton(left,text=skill,variable=teach_vars[skill],bg="#1a1a2e",fg="white",selectcolor="#4a4a8a").pack(anchor="w")
            tk.Checkbutton(right,text=skill,variable=learn_vars[skill],bg="#1a1a2e",fg="white",selectcolor="#4a4a8a").pack(anchor="w")
    def make_cmd(c):
        def cmd():
            show_skills(c)
        return cmd
    for cat in categories:
        tk.Button(cat_frame,text=cat,bg="#4a4a8a",fg="white",command=make_cmd(cat)).pack(side="left",padx=5)
    def save():
        teach_list = []
        learn_list = []

        for skill in teach_vars:
            if teach_vars[skill].get() == True:
                teach_list.append(skill)

        for skill in learn_vars:
            if learn_vars[skill].get() == True:
                learn_list.append(skill)

        teaching = "-".join(teach_list)
        learning = "-".join(learn_list)

        if teaching == "" or learning == "":
            result_label.config(text="Dono fill karo!")
        else:
            save_skills(student_id, teaching, learning)
            result_label.config(text="Saved! ✅")

    result_label = tk.Label(frame, text="",
                            bg="#1a1a2e", fg="white")
    result_label.pack(side="bottom")

    tk.Button(frame, text="Save Skills",
              bg="#4a4a8a", fg="white",
              command=save).pack(side="bottom", pady=10)        
##def skills_ui(student_id, frame):
##    tk.Label(frame, text="Add Your Skills", font=("Arial", 14, "bold"),bg="#1a1a2e",fg="white").pack(pady=10)
##    bottom = tk.Frame(frame,bg="#1a1a2e")
##    bottom.pack(side="bottom",pady=10)
##    result_label = tk.Label(bottom,text="",bg="#1a1a2e",fg="white")
##    result_label.pack()
##    save_btn=tk.Button(bottom,text="Save Skills",bg="#4a4a8a",fg="white")
##    save_btn.pack(pady=5)
##
##    # 2 frames - teach aur learn
##    skill_teach = tk.Frame(frame,bg="#1a1a2e")
##    skill_teach.pack(side="left", padx=80)
##
##    skill_learn = tk.Frame(frame,bg="#1a1a2e")
##    skill_learn.pack(side="right", padx=30)
##
##    skills_list = ["Python", "JAVA", "PHP", "SQL", "C#", "C++", "React", "JavaScript"]
##
##    # Teach section
##    tk.Label(skill_teach, text="Jo Mujhe Aata Hai", font=("Arial", 11, "bold"),bg="#1a1a2e",fg="white").pack()
##    cat1 = tk.StringVar()
##    cat1.set("Select Category")
##    dd = tk.OptionMenu(skill_teach, cat1, *categories.keys())
##    dd.config(bg="#4a4a8a", fg="white")
##    dd.pack()
##
### Skill dropdown
##    skill1 = tk.StringVar()  # ← StringVar banao!
##    skill1.set("Select Skill")
##    skill_dd = tk.OptionMenu(skill_teach, skill1, "Select Category First")
##    skill_dd.config(bg="#4a4a8a", fg="white")
##    skill_dd.pack()
##
### Update function
##    def update_teach(*args):
##        cat = cat1.get()
##        if cat in categories:
##            skills = categories[cat]
##            skill_dd["menu"].delete(0, "end")
##            for skill in skills:
##                def set_skill(s=skill):
##                    skill1.set(s)
##                skill_dd["menu"].add_command(
##                    label=skill,
##                    command=set_skill)
##
##    cat1.trace("w", update_teach)
##
### List aur label
##    teach_skills = []
##    ad_label = tk.Label(skill_teach, text="Added: ",
##                    bg="#1a1a2e", fg="white")
##    ad_label.pack()
##
### Add function
##    def add_teach():
##        skill = skill1.get()
##        if skill != "Select Skill" and skill != "Select Category First":
##            if skill not in teach_skills:
##                teach_skills.append(skill)
##                ad_label.config(text="Added: " + ", ".join(teach_skills))
##            else:
##                ad_label.config(text="Already added!")
##
##    tk.Button(skill_teach, text="Add",
##          bg="#4a4a8a", fg="white",
##          command=add_teach).pack()
##
##    # Learn section
####    categories={"Computer": ["Python","JAVA","PHP","SQL","C#",
####                             "C++","JavaScript","React","HTML","CSS","WordPress",
####                             "PhotoShop","Video Editing"]
####                "Cooking": ["Baking","BBQ","Biryani","Cake Decoration",
####                            "Pizza","Pakistani Dishes","Chinese Food",
####                            "Desserts"]
####                "Arts": ["Drawing","Painting","Calligraphy","Sketching","Digital Art",
####                         "Origami","Pottery"]
####                "Singing": ["Classical","Pop","Nasheed","Rap","Folk","Qawali"]
####                "Languages": ["English","Urdu","Arabic","French","Chinese",
####                              "German"]
####                "Academic": ["Math","Physics","Chemistry","Biology","History",
####                             "Islamiat"]
####                "Fitness": ["Yoga","Gym","Martial Arts","Meditation"]}
##    tk.Label(skill_learn,text="Jo Mujhe Seekhna Hy",font=("Arial",11,"bold"),bg="#1a1a2e",fg="white").pack()            
##    cat2 = tk.StringVar()
##    cat2.set("Select Category")
##    dd2 = tk.OptionMenu(skill_learn, cat2, *categories.keys())
##    dd2.config(bg="#4a4a8a", fg="white")
##    dd2.pack()
##
### Step 2: Skill dropdown
##    skill2 = tk.StringVar()
##    skill2.set("Select Skill")
##    skill_dd2 = tk.OptionMenu(skill_learn, skill2, "Select Category First")
##    skill_dd2.config(bg="#4a4a8a", fg="white")
##    skill_dd2.pack()
##
### Step 3: Update function
##    def update_learn(*args):
##        cat = cat2.get()
##        if cat in categories:
##            skills = categories[cat]
##            skill_dd2["menu"].delete(0, "end")
##            for skill in skills:
##                def set_skill2(s=skill):
##                    skill2.set(s)
##                skill_dd2["menu"].add_command(
##                    label=skill,
##                    command=set_skill2)
##
##    cat2.trace("w", update_learn)
##
### Step 4: List aur label
##    learn_skills = []
##    ad_label2 = tk.Label(skill_learn, text="Added: ",
##                     bg="#1a1a2e", fg="white")
##    ad_label2.pack()
##
### Step 5: Add function
##    def add_learn():
##        skill = skill2.get()
##        if skill != "Select Skill" and skill != "Select Category First":
##            if skill not in learn_skills:
##                learn_skills.append(skill)
##                ad_label2.config(text="Added: " + ", ".join(learn_skills))
##            else:
##                ad_label2.config(text="Already Added!")
##
##    tk.Button(skill_learn, text="Add",
##          bg="#4a4a8a", fg="white",
##          command=add_learn).pack()
##    # Save button
####    result_label = tk.Label(frame, text="",bg="#16213e")
####    result_label.pack(pady=5)
##
##    def save():
##        teaching = "-".join(teach_skills)
##        learning = "-".join(learn_skills)
##        if teaching == "" or learning == "":
##            result_label.config(text="Dono columns fill karo!")
##        else:
##            save_skills(student_id, teaching, learning)
##            result_label.config(text="Skills Saved! ✅")
##            save_btn.config(state="disabled")
##    save_btn.config(command=save)

##    save_btn = tk.Button(frame, text="Save Skills", command=save)
##    save_btn.pack(pady=10)
# Matches: meri learning skills
def find_matches(student_id):
    wb = load_workbook("students.xlsx")
    ws = wb["Skills"]

    learn = []
    for row in ws.iter_rows(values_only=True):
        if row[0] == student_id:
            learn=row[2].split("-")
    # dusre studenet ki teach skill se compare karke common nikalna
    matches = []
    for row in ws.iter_rows(values_only=True):
        if row[0]!=student_id:
            teach = row[1].split("-")
            common=[]
            for skill in learn:
                if skill in teach:
                    common.append(skill)
    # percentage banana
            if len(learn)>0:
                count = len(common)
                total = len(learn)
                percent = count/total*100
                if percent>0:
                    matches.append((row[0],percent,common))
    def percentage(item):
        return item[1]
    matches.sort(key=percentage,reverse = True)
    return matches
def gt_profile(student_id):
    wb = load_workbook("students.xlsx")
    ws1=wb["Sheet1"]
    name = ""
    city=""
    ratings=0
    exchange=0
    for row in ws1.iter_rows(values_only=True):
        if row[0]==student_id:
            name=row[1]
            city=row[2]
            ratings=row[4]
            exhchange=row[5]
    ws2=wb["Skills"]
    teach="not added yet"
    learn = "not added yet"
    for row in ws2.iter_rows(values_only=True):
        if row[0]==student_id:
            teach=row[1]
            learn=row[2]
    return name,city,ratings,exchange,teach,learn
            
    
def gt_dashboard(student_id):
    results = find_matches(student_id)
    matches_count = len(results)

    wb = load_workbook("students.xlsx")
    if "Requests" not in wb.sheetnames:
        return matches_count,0,0
    ws = wb["Requests"]
    sent = 0
    receive = 0
    for row in ws.iter_rows(values_only=True):
        if row[0]==student_id:
            sent+=1
        if row[1]==student_id:
            receive +=1
    return matches_count,sent,receive









    
def accp_req(from_id,to_id):
    wb = load_workbook("students.xlsx")
    ws=wb["Requests"]
    for row in ws.iter_rows():
        if row[0].value==from_id and row[1].value==to_id:
            row[2].value="accepted"
            wb.save("students.xlsx")
            return
def rej_req(from_id,to_id):
    wb = load_workbook("students.xlsx")
    ws = wb["Requests"]
    for row in ws.iter_rows():
        if row[0].value == from_id and row[1].value == to_id:
            row[2].value = "rejected"  # ← status update!
            wb.save("students.xlsx")
            return


        

def sd_message(from_id,to_id,message):
    wb = load_workbook("students.xlsx")
    ws = wb["message"]
    now = datetime.datetime.now()
    time=now.strftime("%H:%M")
    ws.append([from_id,to_id,message,time])
    wb.save("students.xlsx")
def gt_msg(from_id,to_id):
    messages = []
    wb = load_workbook("students.xlsx")
    ws = wb["message"]
    for row in ws.iter_rows(values_only=True):
        if row[0]== from_id and row[1]== to_id:
            messages.append(row)
        elif row[0]==to_id and row[1]==from_id:
            messages.append(row)
    return messages
def gt_connect(student_id):
    conn = []
    wb = load_workbook("students.xlsx")
    ws=wb["Requests"]
    for row in ws.iter_rows(values_only=True):
        if row[0]==student_id and row[2]=="accepted":
            if row[1] not in conn:
                conn.append(row[1])
        elif row[1]==student_id and row[2]=="accepted":
            if row[0] not in conn:
                conn.append(row[0])
    return conn



def gt_stats():
    wb = load_workbook("students.xlsx")
    ws = wb["Skills"]
    skill_count={}
    for row in ws.iter_rows(values_only=True):
        if row[1]:
            skills=row[1].split("-")
            for skill in skills:
                if skill not in skill_count:
                    skill_count[skill]=0
                skill_count[skill]+=1
    return skill_count
##def gt_skill_count():
##    wb = load_workbook("students.xlsx")
##    ws=wb["Skills"]
##    counts={}
##    for row in ws.iter_rows(values_only=True):
##        if row[1]:
##            skills = row[1].split("-")
##            for skill in skills:
##                if skill in counts:
##                    counts[skill]+=1
##                else:
##                    counts[skill]=1
##    return counts
def gt_pie():
    wb = load_workbook("students.xlsx")
    ws = wb["Skills"]
    skill_count={}
    for row in ws.iter_rows(values_only=True):
        if row[1]:
            skills=row[1].split("-")
            for skill in skills:
                if skill not in skill_count:
                    skill_count[skill]=0
                skill_count[skill]+=1
    data = []
    for skill in skill_count:
        count = skill_count[skill]
        data.append((skill,count))
    for i in range(len(data)):
        for j in range(i+1,len(data)):
            if data[j][1] > data[i][1]:
                data[i], data[j] = data[j],data[i]
    top5 = data[:5]
    names=[]
    counts =[]
    for item in top5:
        names.append(item[0])
        counts.append(item[1])
    return names,counts
                         
def gt_req_stats():
    wb = load_workbook("students.xlsx")
    if "Requests" not in wb.sheetnames:
        return 0,0,0
    ws = wb["Requests"]
    pending=0
    accepted = 0
    rejected = 0
    for row in ws.iter_rows(values_only=True):
        if row[2]=="pending":
            pending += 1
        elif row[2]=="accepted":
            accepted += 1
        elif row [2] == "rejected":
            rejected += 1
    return pending,accepted,rejected

def open_main(root, student_id):
    main_win = tk.Toplevel(root)
    main_win.title("SkillSwap")
    main_win.geometry("1000x600")
    main_win.state("zoomed")
    main_win.resizable(False,False)
    main_win.configure(bg="#1a1a2e")

    # Header
    tk.Label(main_win, text="SkillSwap", font=("Arial", 18, "bold"),bg = "#1a1a2e",fg="white").pack(pady=5)
    tk.Label(main_win, text="Learn By Teaching - Teach By Learn", font=("Arial", 10),bg = "#1a1a2e",fg="#a0a0ff").pack()

    # Navbar
    navbar = tk.Frame(main_win, bg="#16213e")
    navbar.pack(pady=5)

    # Content frame
    content = tk.Frame(main_win,bg="#16213e")
    content.pack(fill="both", expand=True) #padx=10, pady=10)

##    nav_img = Image.open(r"C:\Users\Indus\OneDrive\Desktop\files\logo.png.png")
##    nav_img=nav_img.resize((40,40))
##    nav_msk=Image.new("L",(40,40),0)
##    nav_draw=ImageDraw.Draw(nav_msk)
##    nav_draw.ellipse((0,0,40,40),fill=255)
##    nav_img.putalpha(nav_msk)
##    nav_logo=ImageTk.PhotoImage(nav_img)
##    tk.Label(navbar,image=nav_logo,bg="#16213e").pack(side="left",padx=5)
##    main_win.nav_logo=nav_logo
##def gt_stats():
##    wb = load_workbook("students.xlsx")
##    ws = wb["Skills"]
##    skill_count={}
##    for row in ws.iter_rows(values_only=True):
##        if row[1]:
##            skills=row[1].split("-")
##            for skill in skills:
##                if skill not in skill_count:
##                    skill_count[skill]=0
##                skill_count[skill]+=1
##    return skill_count
##def gt_skill_count():
##    wb = load_workbook("students.xlsx")
##    ws=wb["Skills"]
##    counts={}
##    for row in ws.iter_rows(values_only=True):
##        if row[1]:
##            skills = row[1].split("-")
##            for skill in skills:
##                if skill in counts:
##                    counts[skill]+=1
##                else:
##                    counts[skill]=1
##    return counts
##def gt_req_stats():
##    wb = load_workbook("students.xlsx")
##    if "Requests" not in wb.sheetnames:
##        return 0,0,0
##    ws = wb["Requests"]
##    pending=0
##    accepted = 0
##    rejected = 0
##    for row in ws.iter_rows(values_only=True):
##        if row[2]=="pending":
##            pending += 1
##        elif row[2]=="accepted":
##            accepted += 1
##        elif row [2] == "rejected":
##            rejected += 1
##    return pending,accepted,rejected
                       

    def clear():
        for w in content.winfo_children():
            w.destroy()

    
    def home():
        clear()
        card=tk.Frame(content,bg="#16213e",relief="raised",bd=5)
        card.pack(pady=50,padx=50,fill="x")
        tk.Label(card,text="Welcome, " + student_id + "!",bg="#16213e", fg="white",font=("Arial", 13, "bold")).pack(pady=5)

        tk.Label(card, text="Welcome to SkillSwap! - Student Skill Exchange Platform ",bg="#16213e",fg="white", font=("Arial", 14,"bold")).pack(pady=10)
        tk.Label(card, text="PAKISTAN KA PEHLA STUDENT SKILL EXCHANGE ",bg="#16213e",fg="#a0a0ff", font=("Arial", 11)).pack(pady=5)
        tk.Label(card,text="✅ Skills Add Karo     ✅ Match Dhundo        ✅ Exchange Karo!",bg="#16213e", fg="#00ff88",font=("Arial", 11)).pack(pady=15)
##        tk.Label(card,text="✅ Match Dhundo",bg="#16213e", fg="#00ff88",font=("Arial", 11)).pack(anchor="w", padx=20, pady=3)
##        tk.Label(card,text="✅ Exchange Karo!",bg="#16213e", fg="#00ff88",font=("Arial", 11)).pack(anchor="w", padx=20, pady=3)
##        tk.Label(card,text="Learn By Teaching - Teach By Learn!",bg="#16213e", fg="#a0a0ff",font=("Arial", 10, "italic")).pack(pady=10)
        wb = load_workbook("students.xlsx")
        ws = wb["announcements"]
        tk.Label(card,text="📢 Admin Announcements",fg="#ffd700",bg="#16213e",font=("Arial",16,"bold")).pack(pady=15)
        for i in ws.iter_rows(values_only = True, min_row = 2):
            tk.Label(card,text=str(i[1]) + " " + str(i[2]), bg="#16213e",fg="#a0a0ff",font=("Arial",10,"bold")).pack(anchor="e",padx=10)
            tk.Label(card,text="💬"+str(i[0]),bg="#16213e",fg="#FF5733",font=("Arial",12)).pack(anchor="w",padx=10,pady=5)
                       

    def dashboard():
        clear()
        matches_count, sent, received = gt_dashboard(student_id)

        tk.Label(content, text="📊 Dashboard",
             font=("Arial", 16, "bold"),
             bg="#1a1a2e", fg="white").pack(pady=10)

        cards_frame = tk.Frame(content, bg="#1a1a2e")
        cards_frame.pack(pady=10, padx=50, fill="x")

    # 3 alag cards
        def stat_card(parent, title, value, color):
            c = tk.Frame(parent, bg="#16213e",
                     relief="raised", bd=2)
            c.pack(side="left", padx=10,
               pady=5, expand=True, fill="x")
            tk.Label(c, text=title, bg="#16213e",
                 fg="#a0a0ff",
                 font=("Arial", 10)).pack(pady=5)
            tk.Label(c, text=str(value), bg="#16213e",
                 fg=color,
                 font=("Arial", 20, "bold")).pack(pady=5)

        stat_card(cards_frame, "🤝 Matches",
              matches_count, "#00ff88")
        stat_card(cards_frame, "📤 Sent",
              sent, "#a0a0ff")
        stat_card(cards_frame, "📥 Received",
              received, "#a0a0ff")

    # Requests section
        tk.Label(content, text="📬 Pending Requests",
             font=("Arial", 12, "bold"),
             bg="#1a1a2e", fg="white").pack(
             anchor="w", padx=50, pady=10)

        wb = load_workbook("students.xlsx")
        if "Requests" in wb.sheetnames:
            ws = wb["Requests"]
            found = False
            for row in ws.iter_rows(values_only=True):
                if row[1] == student_id and row[2] == "pending":
                    found = True
                    req_frame = tk.Frame(content, bg="#16213e",
                                     relief="raised", bd=1)
                    req_frame.pack(fill="x", padx=50, pady=3)

                    tk.Label(req_frame,
                         text="From: " + str(row[0]),
                         bg="#16213e",
                         fg="white").pack(side="left", padx=10)

                    tk.Label(req_frame,
                         text="Common: " + str(row[3]),
                         bg="#16213e",
                         fg="#a0a0ff").pack(side="left", padx=5)

                    def accept(f=row[0], t=row[1]):
                        accp_req(f, t)
                        dashboard()

                    def reject(f=row[0], t=row[1]):
                        rej_req(f, t)
                        dashboard()

                    tk.Button(req_frame, text="Accept ✅",
                          bg="green", fg="white",
                          command=accept).pack(side="left",
                          padx=5, pady=5)

                    tk.Button(req_frame, text="Reject ❌",
                          bg="red", fg="white",
                          command=reject).pack(side="left",
                          padx=5)

            if not found:
                tk.Label(content,
                     text="Koi pending request nahi!",
                     bg="#1a1a2e", fg="#a0a0ff").pack()
        
        

    def my_skills():
        clear()
        skills_ui(student_id, content)

    def matches():
        clear()
        results = find_matches(student_id)
        tk.Label(content,text="Your Matches",bg="#16213e",fg="white",font=("Arial",14,"bold")).pack()
        if len(results)==0:
            tk.Label(content,text="No Macthes Found").pack()
            return
##        for name,percent in results:
##            row = tk.Frame(content)
##            row.pack(pady=5)
##            tk.Label(row,text=name).pack(side="left",padx=10)
##            tk.Label(row,text=f"{percent:.1f}%").pack(side="left")
        card="#1a1a2e"
        for name,percent,common in results:
            row = tk.Frame(content,bg=card,relief="raised",bd=2)
            row.pack(pady=5,padx=10,fill="x")
            tk.Label(row,text=name,bg=card,fg="white",font=("Arial",11,"bold")).pack(side="left",padx=10)
            tk.Label(row,text=f"{percent:.1f}%",bg=card,fg="#a0a0ff").pack(side="left")
            common_str = ", ".join(common)
            tk.Label(row,text="Common: "+common_str,bg=card,fg="#a0a0ff").pack(side="left",padx=10)
##            row = tk.Frame(content,bg="white",relief="raised",bd=2)
##            row.pack(pady=5,padx=10,fill="x")
##            tk.Button(row,text="Send Request",comm
            def send(n=name,c=common,btn=None):
                result=request(student_id,n,c)
                if result=="sent":
                    btn.config(text="Pending",state="disabled",bg="orange")
                else:
                    btn.config(text="Already Sent!",state="disabled")
            sb=tk.Button(row,text="Send Requests",bg="#4a4a8a",fg="white",relief="flat")#command=send).pack(side="left",padx=10
            def click(b=sb,n=name,c=common):
                send(n,c,b)
            sb.config(command=click)
            sb.pack(side="left",padx=10)
            
##            for row in ws.iter_rows(values_only=True):
##                if row[0]== from_id and row[1] == to_id:
##                    tk.Label(content,text="Already Sent").pack(side="right",pady=5)
                    
            
##            tk.Label(row,text=common).pack()
##        tk.Label(content, text="Matches", font=("Arial", 14, "bold")).pack(pady=20)
##        find_matches(student_id)
##        tk.Label(content, text="Coming soon...").pack()

    def messages():
        selected_conn=tk.StringVar()
        selected_conn.set("")
        clear()
##        tk.Label(content, text="Messages", font=("Arial", 14, "bold")).pack(pady=20)
##        tk.Label(content, text="Coming soon...").pack()
        left = tk.Frame(content,bg="#16213e")
        left.pack(side="left",fill="y",padx=5)
        tk.Label(left,text="Connections",bg="#16213e",fg="white",font=("Arial",11,"bold")).pack(pady=10)
        right = tk.Frame(content,bg="#1a1a2e")
        right.pack(side="right",fill="both",expand=True)
##        tk.Label(right,text="hello2").pack()

        conns = gt_connect(student_id)
        if len(conns)==0:
            tk.Label(left,text="No Connections Yet!",bg="#16213e",fg="white").pack(padx=10)
        for conn in conns:
            def chat(c=conn):
                selected_conn.set(c)
                lc(c)
            tk.Button(left,text=conn,command=chat,bg="#4a4a8a",fg="white",width=15).pack(pady=3,padx=5)
        cbox = tk.Text(right,bg="#16213e",fg="white",state="disabled",height=15,width=40)
        cbox.pack(padx=10,pady=10)
        msg_ent=tk.Entry(right,width=35,bg="#16213e",fg="white",insertbackground="white")
        msg_ent.pack(side="left",padx=5,pady=5)
        def send():
            msg = msg_ent.get()
            if msg !="":
                sd_message(student_id,selected_conn.get(),msg)
                msg_ent.delete(0,"end")
                lc(selected_conn.get())
        tk.Button(right,text="Send",bg="#4a4a8a",fg="white",command=send).pack(side="left",pady=5)
        def lc(other_id):
            msgs = gt_msg(student_id,other_id)
            cbox.config(state="normal")
            cbox.delete("1.0","end")
            for row in msgs:
                cbox.insert("end",row[0]+ ": "+row[2] + " : " + str(row[3])+ "\n")
            cbox.config(state="disabled")
                

    def stats():
        clear()
        plt.close("all")
        main_win.geometry("1000x600")
        tk.Label(content, text="Stats & Analytics", font=("Arial", 14, "bold"),bg="#1a1a2e",fg="white").pack(pady=20)
##        tk.Label(content, text="Coming soon...").pack()
        skill_count = gt_stats()
##        pending,accepted,rejected = gt_req_stats()
        fig,(ax1,ax2) = plt.subplots(1,2)#figsize=(4,2),dpi=70)
        fig.patch.set_facecolor("#1a1a2e")

        ax1.bar(skill_count.keys(),skill_count.values(),color="#4a4a8a")
        ax1.set_title("Skill Count",color="white")
        ax1.tick_params(colors="white",axis="x",rotation=45)
        ax1.tick_params(colors="white",axis="y")
        ax1.set_facecolor("#16213e")
        #graph 2
##        labels = ["Pending","Accepted","Rejeccted"]
##        values=[pending,accepted,rejected]
        names,counts = gt_pie()
        colors=["#ff6b6b","#4ecdc4","#45b7d1","#96ceb4","#ffeaa7"]
        ax2.pie(counts,labels=names,colors=colors,autopct="%1.1f%%",textprops={"color":"white"})
        ax2.set_title("Top 5 Popular Skills",color="white")
        ax2.set_facecolor("#16213e")
        #tkinter
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig,master=content)
        canvas.draw()
        canvas.get_tk_widget().config(width=900,height=400)
        canvas.get_tk_widget().pack(padx=10,pady=5)
        
        

    def profile():
        clear()
    
    # Data pehle lo
        name,city,ratings,exchange,teach,learn = gt_profile(student_id)
        matches, sent, received = gt_dashboard(student_id)
    
    # Title
        tk.Label(content, text="👤 My Profile",
             font=("Arial", 16, "bold"),
             bg="#1a1a2e", fg="white").pack(pady=10)
    
    # Banner
        banner = tk.Frame(content, bg="#0f0f23",
                      relief="raised", bd=2)
        banner.pack(fill="x", padx=50, pady=10)
    
        tk.Label(banner,
             text="👤 " + str(name),
             font=("Arial", 24, "bold"),
             bg="#0f0f23", fg="#00ff88").pack(pady=10)
    
        tk.Label(banner,
             text="🆔 " + student_id +
                  "   🏙️ " + str(city),
             font=("Arial", 11),
             bg="#0f0f23", fg="#a0a0ff").pack()
    
        tk.Label(banner,
             text="⭐ Ratings: " + str(ratings),
             font=("Arial", 11),
             bg="#0f0f23", fg="#ffd700").pack(pady=(0,10))
    
    # Mini cards
        def mini_card(parent, title, value, color):
            c = tk.Frame(parent, bg="#16213e",
                     relief="raised", bd=2)
            c.pack(side="left", padx=15,
               ipadx=20, ipady=10)
            tk.Label(c, text=title,
                 bg="#16213e", fg="#a0a0ff",
                 font=("Arial", 10)).pack()
            tk.Label(c, text=str(value),
                 bg="#16213e", fg=color,
                 font=("Arial", 20, "bold")).pack()
    
        stats_frame = tk.Frame(content, bg="#1a1a2e")
        stats_frame.pack(pady=10)
    
        mini_card(stats_frame, "🤝 Matches",
              matches, "#00ff88")
        mini_card(stats_frame, "📤 Sent",
              sent, "#a0a0ff")
        mini_card(stats_frame, "📥 Received",
              received, "#ffd700")
    
    # Skills card
        skill_card = tk.Frame(content, bg="#16213e",
                          relief="raised", bd=2)
        skill_card.pack(pady=5, padx=50, fill="x")
    
        tk.Label(skill_card, text="My Skills",
             font=("Arial", 11, "bold"),
             bg="#16213e", fg="#a0a0ff").pack(
             anchor="w", padx=20, pady=5)
    
        left_skills = tk.Frame(skill_card, bg="#16213e")
        left_skills.pack(side="left", padx=20, pady=10)
    
        right_skills = tk.Frame(skill_card, bg="#16213e")
        right_skills.pack(side="right", padx=20, pady=10)
    
    # Teaching
        tk.Label(left_skills, text="🎓 Teaching",
             bg="#16213e", fg="#a0a0ff",
             font=("Arial", 11, "bold")).pack()
    
        teach_list = str(teach).split("-")
        for skill in teach_list:
            tk.Label(left_skills,
                 text="• " + skill,
                 bg="#16213e", fg="#00ff88",
                 font=("Arial", 10)).pack(anchor="w")
    
    # Learning
        tk.Label(right_skills, text="📚 Learning",
             bg="#16213e", fg="#a0a0ff",
             font=("Arial", 11, "bold")).pack()
    
        learn_list = str(learn).split("-")
        for skill in learn_list:
            tk.Label(right_skills,
                 text="• " + skill,
                 bg="#16213e", fg="#00ff88",
                 font=("Arial", 10)).pack(anchor="w")
    def logout():
        main_win.destroy()
        root.deiconify()
    def report_issue():
        clear()
        tk.Label(content, text="📢 Report to Admin",
             font=("Arial", 16, "bold"),
             bg="#1a1a2e", fg="#ffd700").pack(pady=15)

        tk.Label(content, text="Apna message admin ko likho:",
             bg="#1a1a2e", fg="#a0a0ff",
             font=("Arial", 11)).pack()

        msg_box = tk.Text(content, height=6, width=50,
                      bg="#16213e", fg="white",
                      font=("Arial", 11),
                      insertbackground="white")
        msg_box.pack(padx=20, pady=10)

        result_lbl = tk.Label(content, text="",
                          bg="#1a1a2e", fg="#00ff88",
                          font=("Arial", 11))
        result_lbl.pack()

        def send_report():
            msg = msg_box.get("1.0", "end").strip()
            if msg == "":
                result_lbl.config(text="Message khali hai!", fg="red")
                return
            wb = load_workbook("students.xlsx")
        # Sheet nahi hai toh banao
            if "admin_reports" not in wb.sheetnames:
                ws = wb.create_sheet("admin_reports")
                ws.append(["student_id", "message", "date", "time"])
            else:
                ws = wb["admin_reports"]
            now = datetime.datetime.now()
            date = now.strftime("%d-%m-%Y")
            time = now.strftime("%H:%M")
            ws.append([student_id, msg, date, time])
            wb.save("students.xlsx")
            msg_box.delete("1.0", "end")
            result_lbl.config(text="✅ Message bhej diya gaya!", fg="#00ff88")

        tk.Button(content, text="📤 Send to Admin",
              bg="#4a4a8a", fg="white",
              font=("Arial", 12), padx=10,
              relief="flat",
              command=send_report).pack(pady=5)
##        tk.Button(navbar,text="Logout",bg="red",fg="white",relief="flat",padx=10,pady=5,command=logout).pack(side="right",padx=2,pady=5)


        

    
    buttons = [
        ("Home", home),
        ("Profile", profile),
        ("Matches", matches),
        ("Dashboard", dashboard),
        ("Messages", messages),
        ("Stats", stats),
        ("Update Skills", my_skills),
        ("📢 Report Issue", report_issue),
        
    ]

    for text, cmd in buttons:
        tk.Button(navbar, text=text,bg="#4a4a8a",fg="white",relief="flat", padx=10,pady=5,command=cmd).pack(side="left",expand=True,fill="x",pady=5)
    tk.Button(navbar,text="Logout",bg="red",fg="white",relief="flat",padx=10,pady=5,command=logout).pack(side="left",expand=True,fill="x",pady=5)


    
    home()

